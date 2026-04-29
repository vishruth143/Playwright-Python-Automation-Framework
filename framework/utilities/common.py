# pylint: disable=[line-too-long, missing-module-docstring, missing-function-docstring, missing-class-docstring]
# pylint: disable=[unused-variable, import-error]

import random
import shutil
from datetime import datetime, timezone

import pandas as pd
from faker import Faker
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import Page

from config.config_parser import ConfigParser
from tests.ui.pta.pages.login_page import LoginPage


class Common:
    """
    Reusable helper methods shared across test suites.

    Playwright equivalent of the Selenium Common class.  Keyboard shortcuts
    (copy/paste/scroll) are handled natively by Playwright via page.keyboard
    and locator.press(), so only the methods actually used in tests are kept.
    """

    # ------------------------------------------------------------------------------------------------------------------
    #                                          Static data-generation helpers
    # ------------------------------------------------------------------------------------------------------------------
    @staticmethod
    def fake_name() -> str:
        return Faker().name()

    @staticmethod
    def fake_first_name() -> str:
        return Faker().first_name()

    @staticmethod
    def fake_last_name() -> str:
        return Faker().last_name()

    @staticmethod
    def fake_ssn() -> str:
        return Faker().ssn().replace('-', '')

    @staticmethod
    def fake_phonenumber() -> str:
        return str(random.randint(2220000000, 2229999999))

    @staticmethod
    def get_date() -> str:
        return datetime.now(timezone.utc).strftime('%m-%d-%Y')

    # ------------------------------------------------------------------------------------------------------------------
    #                                          Instance helpers
    # ------------------------------------------------------------------------------------------------------------------

    def __init__(self, page: Page, test_data=None):
        self.page = page
        self.test_data = test_data
        self.loginpage = LoginPage(self.page)

    def pta_login(self, region: str) -> None:
        """
        Log in to the PTA application using credentials from the env config.

        :param region: Environment key (e.g. "QA", "DEV").
        """
        ui_test_env_config = ConfigParser.load_config("pta_ui_test_env_config")
        env_config = ui_test_env_config.get(region.upper(), {})

        username = env_config.get("username")
        password = env_config.get("password")

        self.loginpage.type_username_input(username)
        self.loginpage.type_password_input(password)
        self.loginpage.click_submit_btn()

    def pta_logout(self) -> None:
        """Log out of the PTA application."""
        self.loginpage.click_logout_btn()


def save_excel(sheet_name: str, df: pd.DataFrame, input_path: str, output_path: str) -> None:
    """
    Copy the source Excel file to output_path, write the updated DataFrame into
    the specified sheet, then colour-code the 'Status' column (Pass=green, Fail=red).

    :param sheet_name: Name of the sheet to write to.
    :param df:         DataFrame containing the updated data.
    :param input_path: Absolute path of the source Excel file.
    :param output_path: Absolute path of the destination Excel file.
    """
    shutil.copy(input_path, output_path)

    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)
    ws = wb[sheet_name]

    green_fill = PatternFill(fill_type="solid", fgColor="90EE90")
    red_fill = PatternFill(fill_type="solid", fgColor="FF7F7F")

    header_values = [cell.value for cell in ws[1]]
    if "Status" in header_values:
        status_col_index = header_values.index("Status") + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell = row[status_col_index - 1]
            if cell.value == 'Pass':
                cell.fill = green_fill
            elif cell.value == 'Fail':
                cell.fill = red_fill

    wb.save(output_path)

